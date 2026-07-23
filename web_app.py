import os
import sys
import argparse
import logging
import webbrowser
import threading
from flask import Flask, render_template, request, redirect, url_for, flash, send_file, jsonify
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from db import *
import sqlite3
import re
from datetime import datetime, timedelta
from pathlib import Path
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    HAVE_OPENPYXL = True
except Exception:
    Workbook = None
    load_workbook = None
    Font = PatternFill = Alignment = None
    HAVE_OPENPYXL = False
from io import BytesIO
try:
    import pandas as pd
    HAVE_PANDAS = True
except Exception:
    pd = None
    HAVE_PANDAS = False
import base64

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(name)s: %(message)s',
    handlers=[logging.StreamHandler()]
)
log = logging.getLogger(__name__)

app = Flask(__name__)
app.secret_key = os.environ.get('FLASK_SECRET_KEY') or os.urandom(24).hex()
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024
try:
    from flask_wtf.csrf import CSRFProtect
    csrf = CSRFProtect(app)
except ImportError:
    csrf = None
    log.warning("Flask-WTF not installed; CSRF protection disabled")

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message_category = 'error'

SEARCH_RATE_LIMIT = 20
SEARCH_RATE_PERIOD = 10
search_request_log = {}
SEARCH_API_KEY = os.environ.get('EQUIPMENT_SEARCH_API_KEY')


class User(UserMixin):
    def __init__(self, id, username, is_admin):
        self.id = id
        self.username = username
        self.is_admin = is_admin


@login_manager.user_loader
def load_user(user_id):
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute("SELECT id, username, is_admin FROM users WHERE id = ?", (int(user_id),))
    row = cursor.fetchone()
    conn.close()
    if row:
        return User(row['id'], row['username'], row['is_admin'])
    return None


@app.context_processor
def inject_is_frozen():
    return dict(is_frozen=getattr(sys, 'frozen', False))


@app.route('/shutdown')
@login_required
def shutdown():
    if not getattr(sys, 'frozen', False):
        return "Shutdown is only available in the standalone application.", 403
    log.info("Shutdown requested via /shutdown route")
    threading.Thread(target=lambda: os._exit(0), daemon=True).start()
    return "Shutting down..."


@app.before_request
def check_first_run():
    if request.endpoint in ('login', 'register', 'static', None):
        return
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) AS cnt FROM users")
    row = cursor.fetchone()
    conn.close()
    if row['cnt'] == 0:
        return redirect(url_for('register'))


@app.after_request
def set_security_headers(response):
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    response.headers['Content-Security-Policy'] = "default-src 'self'; script-src 'self' 'unsafe-inline'; style-src 'self' 'unsafe-inline'; img-src 'self' data:"
    return response


@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('master_control'))
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) AS cnt FROM users")
    row = cursor.fetchone()
    conn.close()
    if row['cnt'] == 0:
        return redirect(url_for('register'))
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("SELECT id, username, password_hash, is_admin FROM users WHERE username = ?", (username,))
        user_row = cursor.fetchone()
        conn.close()
        if user_row and check_password_hash(user_row['password_hash'], password):
            user = User(user_row['id'], user_row['username'], user_row['is_admin'])
            login_user(user)
            flash('Logged in successfully.')
            next_page = request.args.get('next')
            return redirect(next_page or url_for('master_control'))
        flash('Invalid username or password.')
    return render_template('login.html')


@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('You have been logged out.')
    return redirect(url_for('login'))


@app.route('/register', methods=['GET', 'POST'])
def register():
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) AS cnt FROM users")
    row = cursor.fetchone()
    user_count = row['cnt']
    conn.close()
    is_public = user_count == 0
    if not is_public and (not current_user.is_authenticated or not current_user.is_admin):
        flash('Only administrators can register new users.')
        return redirect(url_for('master_control'))
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        confirm = request.form.get('confirm_password', '')
        if not username or not password:
            flash('Username and password are required.')
            return redirect(url_for('register'))
        if len(username) < 3:
            flash('Username must be at least 3 characters.')
            return redirect(url_for('register'))
        if len(password) < 8:
            flash('Password must be at least 8 characters.')
            return redirect(url_for('register'))
        if password != confirm:
            flash('Passwords do not match.')
            return redirect(url_for('register'))
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM users WHERE username = ?", (username,))
        if cursor.fetchone():
            conn.close()
            flash('Username already exists.')
            return redirect(url_for('register'))
        is_admin = 1 if user_count == 0 else 0
        cursor.execute(
            "INSERT INTO users (username, password_hash, is_admin) VALUES (?, ?, ?)",
            (username, generate_password_hash(password), is_admin),
        )
        conn.commit()
        conn.close()
        flash('Account created successfully. Please log in.')
        return redirect(url_for('login'))
    return render_template('register.html')


@app.route('/change_password', methods=['GET', 'POST'])
@login_required
def change_password():
    if request.method == 'POST':
        current_password = request.form.get('current_password', '')
        new_password = request.form.get('new_password', '')
        confirm = request.form.get('confirm_password', '')
        if not current_password or not new_password:
            flash('All fields are required.')
            return redirect(url_for('change_password'))
        if new_password != confirm:
            flash('New passwords do not match.')
            return redirect(url_for('change_password'))
        if len(new_password) < 8:
            flash('New password must be at least 8 characters.')
            return redirect(url_for('change_password'))
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("SELECT password_hash FROM users WHERE id = ?", (current_user.id,))
        row = cursor.fetchone()
        if not row or not check_password_hash(row['password_hash'], current_password):
            conn.close()
            flash('Current password is incorrect.')
            return redirect(url_for('change_password'))
        cursor.execute(
            "UPDATE users SET password_hash = ? WHERE id = ?",
            (generate_password_hash(new_password), current_user.id),
        )
        conn.commit()
        conn.close()
        flash('Password changed successfully.')
        return redirect(url_for('master_control'))
    return render_template('change_password.html')


@app.route('/')
def index():
    return redirect(url_for('master_control'))


@app.route('/master', methods=['GET', 'POST'])
@login_required
def master_control():
    checkout_candidates = None
    pending_equipment_ids = []
    pending_customer_reference = None

    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'add_customer':
            name = request.form['name'].strip()
            phone = request.form['phone'].strip()
            zip_code = request.form['zip_code'].strip()

            if not name or not zip_code:
                flash('Name and ZIP Code are required.')
                return redirect(url_for('master_control'))

            digits = normalize_phone(phone) if phone else ''
            if phone and len(digits) < 10:
                flash('Phone number must have at least 10 digits.')
                return redirect(url_for('master_control'))

            if not re.fullmatch(r"\d{5}", zip_code):
                flash('Zip code must be 5 digits.')
                return redirect(url_for('master_control'))

            conn = connect_db()
            cursor = conn.cursor()

            if digits and customer_phone_exists(digits, cursor):
                conn.close()
                flash('A customer with this phone number already exists.')
                return redirect(url_for('master_control'))

            formatted_phone = format_phone(phone) if phone else ''
            try:
                cursor.execute(
                    "INSERT INTO customers (name, phone, zip_code, date_added) VALUES (?, ?, ?, ?)",
                    (name, formatted_phone, zip_code, datetime.today().date().isoformat()),
                )
                conn.commit()
                flash(f'Customer "{name}" added successfully.')
            except sqlite3.IntegrityError:
                flash('Error adding customer.')
            finally:
                conn.close()

        elif action == 'add_equipment':
            equipment_id = request.form['equipment_id'].strip().upper()
            item_name = request.form['item_name'].strip()

            if not equipment_id or not item_name:
                flash('All equipment fields are required.')
                return redirect(url_for('master_control'))

            if not EQUIPMENT_ID_PATTERN.fullmatch(equipment_id):
                flash('Equipment ID must be in format AA-0000.')
                return redirect(url_for('master_control'))

            conn = connect_db()
            cursor = conn.cursor()
            try:
                cursor.execute(
                    "INSERT INTO equipment (equipment_id, item_name) VALUES (?, ?)",
                    (equipment_id, item_name),
                )
                conn.commit()
                flash(f'Equipment "{equipment_id}" added successfully.')
            except sqlite3.IntegrityError:
                flash('Equipment ID already exists.')
            finally:
                conn.close()

        elif action == 'checkout':
            customer_reference = request.form.get('customer_reference', '').strip()
            equipment_ids = request.form.getlist('equipment_ids')
            if not equipment_ids:
                equipment_id = request.form.get('equipment_id', '').strip().upper()
                if equipment_id:
                    equipment_ids = [equipment_id]
            equipment_ids = [eid.strip().upper() for eid in equipment_ids if eid.strip()]

            if not equipment_ids:
                flash('Please select at least one piece of equipment.')
                return redirect(url_for('master_control'))

            if not customer_reference:
                conn = connect_db()
                cursor = conn.cursor()
                checked_out_date = datetime.today().date()
                due_date = checked_out_date + timedelta(days=CHECKOUT_PERIOD_DAYS)
                cursor.execute(
                    "INSERT INTO customers (name, phone, zip_code, date_added) VALUES (?, ?, ?, ?)",
                    ('Unknown', '', '00000', checked_out_date.isoformat()),
                )
                new_customer_id = cursor.lastrowid
                loan_ids = []
                for equipment_id in equipment_ids:
                    cursor.execute(
                        "SELECT equipment_id, item_name FROM equipment WHERE equipment_id = ?",
                        (equipment_id,)
                    )
                    equipment_row = cursor.fetchone()
                    if not equipment_row:
                        conn.close()
                        flash(f'Equipment {equipment_id} does not exist.')
                        return redirect(url_for('master_control'))
                    cursor.execute(
                        "SELECT id FROM loans WHERE equipment_id = ? AND returned_date IS NULL",
                        (equipment_id,)
                    )
                    if cursor.fetchone():
                        conn.close()
                        flash(f'Equipment {equipment_id} is already checked out.')
                        return redirect(url_for('master_control'))
                    cursor.execute(
                        "INSERT INTO loans (customer_id, equipment_id, checked_out_date, due_date, agreement_data, agreement_date, agreement_pending) VALUES (?, ?, ?, ?, ?, ?, 1)",
                        (new_customer_id, equipment_id, checked_out_date.isoformat(), due_date.isoformat(), None, None),
                    )
                    loan_id = cursor.lastrowid
                    loan_ids.append(str(loan_id))
                conn.commit()
                conn.close()
                return redirect(url_for('customer_agreement', customer_id=new_customer_id, loan_ids=','.join(loan_ids), new_customer=1))

            matches = find_customer_matches(customer_reference)
            if not matches:
                flash('Customer not found. Use an existing ID, name, phone, or ZIP.')
                return redirect(url_for('master_control'))

            if len(matches) > 1:
                pending_equipment_ids = equipment_ids
                pending_customer_reference = customer_reference
                checkout_candidates = matches
            else:
                customer_id = matches[0]["id"]
                customer_zip = matches[0]["zip_code"]
                conn = connect_db()
                cursor = conn.cursor()
                checked_out_date = datetime.today().date()
                due_date = checked_out_date + timedelta(days=CHECKOUT_PERIOD_DAYS)
                loan_ids = []
                for equipment_id in equipment_ids:
                    cursor.execute(
                        "SELECT equipment_id, item_name FROM equipment WHERE equipment_id = ?",
                        (equipment_id,)
                    )
                    equipment_row = cursor.fetchone()
                    if not equipment_row:
                        conn.close()
                        flash(f'Equipment {equipment_id} does not exist.')
                        return redirect(url_for('master_control'))

                    cursor.execute(
                        "SELECT id FROM loans WHERE equipment_id = ? AND returned_date IS NULL",
                        (equipment_id,)
                    )
                    if cursor.fetchone():
                        conn.close()
                        flash(f'Equipment {equipment_id} is already checked out.')
                        return redirect(url_for('master_control'))

                    cursor.execute(
                        "INSERT INTO loans (customer_id, equipment_id, checked_out_date, due_date, agreement_data, agreement_date, agreement_pending) VALUES (?, ?, ?, ?, ?, ?, 1)",
                        (customer_id, equipment_id, checked_out_date.isoformat(), due_date.isoformat(), None, None),
                    )
                    loan_id = cursor.lastrowid
                    loan_ids.append(str(loan_id))
                conn.commit()
                conn.close()
                return redirect(url_for('customer_agreement', customer_id=customer_id, loan_ids=','.join(loan_ids)))

        elif action == 'return':
            loan_id = request.form['loan_id']
            if loan_id:
                returned_date = datetime.today().date().isoformat()
                conn = connect_db()
                cursor = conn.cursor()
                cursor.execute(
                    "UPDATE loans SET returned_date = ? WHERE id = ?",
                    (returned_date, loan_id),
                )
                conn.commit()
                conn.close()
                flash('Equipment returned successfully.')

        if checkout_candidates:
            pass
        else:
            return redirect(url_for('master_control', search=request.args.get('search', ''), sort_by=request.args.get('sort_by', 'equipment_id'), sort_dir=request.args.get('sort_dir', 'asc'), date_from=request.args.get('date_from', ''), date_to=request.args.get('date_to', '')))

    search = request.args.get('search', '').strip()
    sort_by = request.args.get('sort_by', 'equipment_id')
    sort_dir = request.args.get('sort_dir', 'asc')
    date_from = request.args.get('date_from', '').strip()
    date_to = request.args.get('date_to', '').strip()
    sort_options = {
        'equipment_id': 'equipment.equipment_id',
        'item_name': 'equipment.item_name',
        'customer_name': 'customers.name',
        'customer_phone': 'customers.phone',
        'checked_out_date': 'loans.checked_out_date',
        'due_date': 'loans.due_date',
    }
    sort_column = sort_options.get(sort_by, 'equipment.equipment_id')
    sort_direction = 'DESC' if sort_dir == 'desc' else 'ASC'

    conn = connect_db()
    cursor = conn.cursor()
    query = (
        "SELECT equipment.equipment_id, equipment.item_name, customers.id AS customer_id, customers.name AS customer_name, "
        "customers.phone AS customer_phone, loans.id AS loan_id, loans.checked_out_date, loans.due_date, loans.agreement_data "
        "FROM equipment "
        "LEFT JOIN loans ON equipment.equipment_id = loans.equipment_id AND loans.returned_date IS NULL AND loans.agreement_pending = 0 "
    )
    params = list()
    if date_from:
        query += "AND loans.checked_out_date >= ? "
        params.append(date_from)
    if date_to:
        query += "AND loans.checked_out_date <= ? "
        params.append(date_to)
    query += "LEFT JOIN customers ON loans.customer_id = customers.id "
    if search:
        search_pattern = f"%{escape_like(search)}%"
        query += (
            "WHERE equipment.equipment_id LIKE ? ESCAPE '\\' OR equipment.item_name LIKE ? ESCAPE '\\' "
            "OR customers.name LIKE ? ESCAPE '\\' OR customers.phone LIKE ? ESCAPE '\\' "
        )
        params.extend([search_pattern, search_pattern, search_pattern, search_pattern])
    query += f"ORDER BY {sort_column} {sort_direction}"
    cursor.execute(query, params)
    rows = cursor.fetchall()

    cursor.execute(
        "SELECT equipment.equipment_id, equipment.item_name FROM equipment "
        "LEFT JOIN loans ON equipment.equipment_id = loans.equipment_id AND loans.returned_date IS NULL "
        "WHERE loans.id IS NULL ORDER BY equipment.equipment_id"
    )
    available_equipment = cursor.fetchall()
    conn.close()

    today_str = datetime.today().date().isoformat()

    return render_template(
        'master.html',
        rows=rows,
        search=search,
        sort_by=sort_by,
        sort_dir=sort_dir,
        date_from=date_from,
        date_to=date_to,
        today_str=today_str,
        checkout_candidates=checkout_candidates,
        pending_equipment_ids=pending_equipment_ids,
        pending_customer_reference=pending_customer_reference,
        available_equipment=available_equipment,
    )


@app.route('/customers')
@login_required
def customers():
    search = request.args.get('search', '').strip()
    conn = connect_db()
    cursor = conn.cursor()
    customer_query = (
        "SELECT customers.id, customers.name, customers.phone, customers.zip_code, customers.date_added, "
        "EXISTS(SELECT 1 FROM loans WHERE loans.customer_id = customers.id "
        "AND loans.returned_date IS NULL AND loans.agreement_data IS NOT NULL) AS has_agreement "
        "FROM customers "
    )
    if search:
        search_pattern = f"%{escape_like(search)}%"
        cursor.execute(
            customer_query +
            "WHERE name LIKE ? ESCAPE '\\' OR phone LIKE ? ESCAPE '\\' OR zip_code LIKE ? ESCAPE '\\' "
            "ORDER BY name",
            (search_pattern, search_pattern, search_pattern),
        )
    else:
        cursor.execute(customer_query + "ORDER BY name")
    customers_list = cursor.fetchall()
    conn.close()
    return render_template('customers.html', customers=customers_list, search=search)


@app.route('/add_customer', methods=['GET', 'POST'])
@login_required
def add_customer():
    if request.method == 'POST':
        name = request.form['name'].strip()
        phone = request.form['phone'].strip()
        zip_code = request.form['zip_code'].strip()

        if not name or not phone or not zip_code:
            flash('All fields are required.')
            return redirect(url_for('add_customer'))

        digits = normalize_phone(phone)
        if len(digits) < 10:
            flash('Phone number must have at least 10 digits.')
            return redirect(url_for('add_customer'))

        if not re.fullmatch(r"\d{5}", zip_code):
            flash('Zip code must be 5 digits.')
            return redirect(url_for('add_customer'))

        conn = connect_db()
        cursor = conn.cursor()

        if customer_phone_exists(digits, cursor):
            conn.close()
            flash('A customer with this phone number already exists.')
            return redirect(url_for('add_customer'))

        formatted_phone = format_phone(phone)
        try:
            cursor.execute(
                "INSERT INTO customers (name, phone, zip_code, date_added) VALUES (?, ?, ?, ?)",
                (name, formatted_phone, zip_code, datetime.today().date().isoformat()),
            )
            conn.commit()
            flash(f'Customer "{name}" added successfully.')
        except sqlite3.IntegrityError:
            flash('Error adding customer.')
        finally:
            conn.close()

        return redirect(url_for('customers'))

    return render_template('add_customer.html')


@app.route('/equipment')
@login_required
def equipment():
    search = request.args.get('search', '').strip()
    conn = connect_db()
    cursor = conn.cursor()
    query = (
        "SELECT equipment.equipment_id, equipment.item_name, loans.id AS loan_id, "
        "loans.checked_out_date, loans.due_date, customers.name AS customer_name "
        "FROM equipment "
        "LEFT JOIN loans ON equipment.equipment_id = loans.equipment_id AND loans.returned_date IS NULL AND loans.agreement_pending = 0 "
        "LEFT JOIN customers ON loans.customer_id = customers.id "
    )
    params = ()
    if search:
        search_pattern = f"%{escape_like(search)}%"
        query += (
            "WHERE equipment.equipment_id LIKE ? ESCAPE '\\' OR equipment.item_name LIKE ? ESCAPE '\\' "
            "OR customers.name LIKE ? ESCAPE '\\' "
        )
        params = (search_pattern, search_pattern, search_pattern)
    query += "ORDER BY equipment.equipment_id"
    cursor.execute(query, params)
    equipment_list = cursor.fetchall()
    conn.close()
    return render_template('equipment.html', equipment=equipment_list, search=search)


@app.route('/add_equipment', methods=['GET', 'POST'])
@login_required
def add_equipment():
    if request.method == 'POST':
        equipment_id = request.form['equipment_id'].strip().upper()
        item_name = request.form['item_name'].strip()

        if not equipment_id or not item_name:
            flash('All fields are required.')
            return redirect(url_for('add_equipment'))

        if not EQUIPMENT_ID_PATTERN.fullmatch(equipment_id):
            flash('Equipment ID must be in format AA-0000.')
            return redirect(url_for('add_equipment'))

        conn = connect_db()
        cursor = conn.cursor()
        try:
            cursor.execute(
                "INSERT INTO equipment (equipment_id, item_name) VALUES (?, ?)",
                (equipment_id, item_name),
            )
            conn.commit()
            flash(f'Equipment "{equipment_id}" added successfully.')
        except sqlite3.IntegrityError:
            flash('Equipment ID already exists.')
        finally:
            conn.close()

        return redirect(url_for('equipment'))

    return render_template('add_equipment.html')


@app.route('/delete_customer/<int:customer_id>', methods=['POST'])
@login_required
def delete_customer(customer_id):
    search = request.form.get('search', '').strip()
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT id FROM loans WHERE customer_id = ? AND returned_date IS NULL",
        (customer_id,)
    )
    if cursor.fetchone():
        conn.close()
        flash('Cannot delete customer while they have active checked out equipment.')
        return redirect(url_for('customers'))

    cursor.execute("DELETE FROM customers WHERE id = ?", (customer_id,))
    conn.commit()
    conn.close()
    flash('Customer deleted successfully.')
    return redirect(url_for('customers'))


@app.route('/delete_equipment/<equipment_id>', methods=['POST'])
@login_required
def delete_equipment(equipment_id):
    search = request.form.get('search', '').strip()
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT id FROM loans WHERE equipment_id = ? AND returned_date IS NULL",
        (equipment_id,)
    )
    if cursor.fetchone():
        conn.close()
        flash('Cannot delete equipment while it is checked out.')
        return redirect(url_for('equipment', search=search))

    cursor.execute("SELECT item_name FROM equipment WHERE equipment_id = ?", (equipment_id,))
    row = cursor.fetchone()
    item_name = row["item_name"] if row else equipment_id

    cursor.execute(
        "INSERT INTO deleted_items_log (equipment_id, item_name, deletion_date) VALUES (?, ?, ?)",
        (equipment_id, item_name, datetime.today().date().isoformat()),
    )

    cursor.execute("DELETE FROM equipment WHERE equipment_id = ?", (equipment_id,))
    conn.commit()
    conn.close()
    flash('Equipment deleted successfully.')
    return redirect(url_for('equipment', search=search))


@app.route('/settings', methods=['GET', 'POST'])
@login_required
def settings():
    if request.method == 'POST':
        action = request.form.get('action')

        if action == 'import_customers':
            if 'file' not in request.files:
                flash('No file part')
                return redirect(url_for('settings'))
            file = request.files['file']
            if file.filename == '':
                flash('No selected file')
                return redirect(url_for('settings'))
            if file and (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
                try:
                    conn = connect_db()
                    cursor = conn.cursor()
                    count = 0
                    if file.filename.endswith('.xls'):
                        if not HAVE_PANDAS:
                            flash('Pandas is not available on the server; cannot import .xls files. Install pandas.')
                        else:
                            df = pd.read_excel(file)
                            for _, row in df.iterrows():
                                try:
                                    name = row.iloc[0]
                                    phone = row.iloc[1]
                                    zip_code = row.iloc[2]
                                except Exception:
                                    continue
                                if name and phone and zip_code:
                                    digits = normalize_phone(str(phone))
                                    if len(digits) >= 10 and re.fullmatch(r"\d{5}", str(zip_code)):
                                        if not customer_phone_exists(digits, cursor):
                                            cursor.execute(
                                                "INSERT INTO customers (name, phone, zip_code, date_added) VALUES (?, ?, ?, ?)",
                                                (str(name).strip(), format_phone(phone), str(zip_code), datetime.today().date().isoformat()),
                                            )
                                            count += 1
                    else:
                        wb = load_workbook(file)
                        ws = wb.active
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            if row[0] and row[1] and row[2]:
                                name, phone, zip_code = row[0], row[1], row[2]
                                digits = normalize_phone(str(phone))
                                if len(digits) >= 10 and re.fullmatch(r"\d{5}", str(zip_code)):
                                    if not customer_phone_exists(digits, cursor):
                                        cursor.execute(
                                            "INSERT INTO customers (name, phone, zip_code, date_added) VALUES (?, ?, ?, ?)",
                                            (str(name).strip(), format_phone(phone), str(zip_code), datetime.today().date().isoformat()),
                                        )
                                        count += 1
                    conn.commit()
                    conn.close()
                    flash(f'Imported {count} new customers successfully.')
                except Exception as e:
                    log.exception("Error importing file")
                    flash('An error occurred while importing the file. Please check the format and try again.')
            else:
                flash('File must be .xlsx or .xls format')
            return redirect(url_for('settings'))

        elif action == 'import_equipment':
            if 'file' not in request.files:
                flash('No file part')
                return redirect(url_for('settings'))
            file = request.files['file']
            if file.filename == '':
                flash('No selected file')
                return redirect(url_for('settings'))
            if file and (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
                try:
                    conn = connect_db()
                    cursor = conn.cursor()
                    count = 0
                    if file.filename.endswith('.xls'):
                        if not HAVE_PANDAS:
                            flash('Pandas is not available on the server; cannot import .xls files. Install pandas.')
                        else:
                            df = pd.read_excel(file)
                            for _, row in df.iterrows():
                                try:
                                    equipment_id = str(row.iloc[0]).upper()
                                    item_name = str(row.iloc[1])
                                except Exception:
                                    continue
                                if equipment_id and item_name and EQUIPMENT_ID_PATTERN.fullmatch(equipment_id):
                                    cursor.execute(
                                        "SELECT id FROM equipment WHERE equipment_id = ?",
                                        (equipment_id,),
                                    )
                                    if not cursor.fetchone():
                                        cursor.execute(
                                            "INSERT INTO equipment (equipment_id, item_name) VALUES (?, ?)",
                                            (equipment_id, item_name),
                                        )
                                        count += 1
                    else:
                        wb = load_workbook(file)
                        ws = wb.active
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            if row[0] and row[1]:
                                equipment_id, item_name = str(row[0]).upper(), str(row[1])
                                if EQUIPMENT_ID_PATTERN.fullmatch(equipment_id):
                                    cursor.execute(
                                        "SELECT id FROM equipment WHERE equipment_id = ?",
                                        (equipment_id,),
                                    )
                                    if not cursor.fetchone():
                                        cursor.execute(
                                            "INSERT INTO equipment (equipment_id, item_name) VALUES (?, ?)",
                                            (equipment_id, item_name),
                                        )
                                        count += 1
                    conn.commit()
                    conn.close()
                    flash(f'Imported {count} new equipment items successfully.')
                except Exception as e:
                    log.exception("Error importing file")
                    flash('An error occurred while importing the file. Please check the format and try again.')
            else:
                flash('File must be .xlsx or .xls format')
            return redirect(url_for('settings'))

        elif action == 'export':
            export_type = request.form.get('export_type')
            conn = connect_db()
            cursor = conn.cursor()

            if export_type == 'customers':
                cursor.execute("SELECT id, name, phone, zip_code, date_added FROM customers ORDER BY id")
                rows = cursor.fetchall()
                conn.close()
                if HAVE_OPENPYXL:
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Customers"
                    ws.append(['ID', 'Name', 'Phone', 'ZipCode', 'DateAdded'])
                    for row in rows:
                        ws.append([row['id'], row['name'], row['phone'], row['zip_code'], row['date_added']])
                    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    header_font = Font(bold=True, color="FFFFFF")
                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                    output = BytesIO()
                    wb.save(output)
                    output.seek(0)
                    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='customers_export.xlsx')
                else:
                    import csv
                    output = BytesIO()
                    writer = csv.writer(output)
                    writer.writerow(['ID', 'Name', 'Phone', 'ZipCode', 'DateAdded'])
                    for row in rows:
                        writer.writerow([row['id'], row['name'], row['phone'], row['zip_code'], row['date_added']])
                    output.seek(0)
                    return send_file(output, mimetype='text/csv', as_attachment=True, download_name='customers_export.csv')

            elif export_type == 'equipment':
                cursor.execute("SELECT equipment_id, item_name FROM equipment ORDER BY equipment_id")
                rows = cursor.fetchall()
                conn.close()
                if HAVE_OPENPYXL:
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Equipment"
                    ws.append(['EquipmentID', 'ItemName'])
                    for row in rows:
                        ws.append([row['equipment_id'], row['item_name']])
                    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    header_font = Font(bold=True, color="FFFFFF")
                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                    output = BytesIO()
                    wb.save(output)
                    output.seek(0)
                    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='equipment_export.xlsx')
                else:
                    import csv
                    output = BytesIO()
                    writer = csv.writer(output)
                    writer.writerow(['EquipmentID', 'ItemName'])
                    for row in rows:
                        writer.writerow([row['equipment_id'], row['item_name']])
                    output.seek(0)
                    return send_file(output, mimetype='text/csv', as_attachment=True, download_name='equipment_export.csv')

    return render_template('settings.html')


@app.route('/reports', methods=['GET', 'POST'])
@login_required
def reports():
    report_type = request.args.get('report_type', 'analytics')
    year_filter = request.args.get('year_filter', '')
    month_filter = request.args.get('month_filter', '')
    date_from = request.args.get('date_from', '').strip()
    date_to = request.args.get('date_to', '').strip()

    if request.method == 'POST':
        action = request.form.get('action')
        report_type = request.form.get('report_type', report_type)
        year_filter = request.form.get('year_filter', year_filter)
        month_filter = request.form.get('month_filter', month_filter)
        date_from = request.form.get('date_from', date_from).strip()
        date_to = request.form.get('date_to', date_to).strip()
        conn = connect_db()
        cursor = conn.cursor()
        if action == 'delete_checkout':
            checkout_id = request.form.get('checkout_id')
            if checkout_id and checkout_id.isdigit():
                cursor.execute("DELETE FROM checkout_log WHERE id = ?", (int(checkout_id),))
                conn.commit()
                flash('Checkout log entry removed successfully.')
        elif action == 'delete_item_sale':
            item_sale_id = request.form.get('item_sale_id')
            if item_sale_id and item_sale_id.isdigit():
                cursor.execute("DELETE FROM deleted_items_log WHERE id = ?", (int(item_sale_id),))
                conn.commit()
                flash('Item sale log entry removed successfully.')
        conn.close()
        return redirect(url_for('reports', report_type=report_type, year_filter=year_filter, month_filter=month_filter, date_from=date_from, date_to=date_to))

    conn = connect_db()
    cursor = conn.cursor()

    report_data = []
    report_title = ''
    analytics_summary = None
    daily_guests = []
    monthly_stats = []
    analytics_months = []

    if report_type == 'checkout':
        query = "SELECT id, customer_zip_code, item_name, equipment_id, checkout_date, is_first_item FROM checkout_log WHERE 1=1"
        params = list()
        if year_filter:
            query += " AND strftime('%Y', checkout_date) = ?"
            params.append(year_filter)
        if date_from:
            query += " AND checkout_date >= ?"
            params.append(date_from)
        if date_to:
            query += " AND checkout_date <= ?"
            params.append(date_to)
        query += " ORDER BY checkout_date DESC"
        cursor.execute(query, params)
        report_data = cursor.fetchall()
        report_title = "Checkout Log"

    elif report_type == 'item_sales':
        query = "SELECT id, equipment_id, item_name, deletion_date FROM deleted_items_log WHERE 1=1"
        params = list()
        if year_filter:
            query += " AND strftime('%Y', deletion_date) = ?"
            params.append(year_filter)
        if date_from:
            query += " AND deletion_date >= ?"
            params.append(date_from)
        if date_to:
            query += " AND deletion_date <= ?"
            params.append(date_to)
        query += " ORDER BY deletion_date DESC"
        cursor.execute(query, params)
        report_data = cursor.fetchall()
        report_title = "Item Sales Log"

    elif report_type == 'analytics':
        report_title = "Analytics"
        year_clause = "AND strftime('%Y', checked_out_date) = ?" if year_filter else ""
        params = (year_filter,) if year_filter else ()

        cursor.execute(
            f"SELECT DISTINCT strftime('%Y-%m', checked_out_date) AS month FROM loans WHERE agreement_pending = 0 {year_clause} ORDER BY month DESC",
            params,
        )
        analytics_months = [row['month'] for row in cursor.fetchall() if row['month']]
        if month_filter not in analytics_months:
            month_filter = analytics_months[0] if analytics_months else ''

        daily_clause = year_clause
        daily_params = params
        if month_filter:
            daily_clause += " AND strftime('%Y-%m', checked_out_date) = ?"
            daily_params += (month_filter,)

        cursor.execute(
            f"""SELECT checked_out_date AS date,
                       COUNT(DISTINCT customer_id || '_' || checked_out_date) AS guest_count,
                       COUNT(*) AS item_count
               FROM loans
               WHERE agreement_pending = 0 {daily_clause}
               GROUP BY checked_out_date
               ORDER BY checked_out_date DESC""",
            daily_params,
        )
        daily_guests = cursor.fetchall()

        cursor.execute(
            f"""SELECT strftime('%Y-%m', checked_out_date) AS month_year,
                       COUNT(DISTINCT customer_id || '_' || checked_out_date) AS total_guests,
                       COUNT(*) AS total_items,
                       COUNT(DISTINCT checked_out_date) AS active_days
               FROM loans
               WHERE agreement_pending = 0 {year_clause}
               GROUP BY month_year
               ORDER BY month_year DESC""",
            params,
        )
        raw_monthly = cursor.fetchall()
        monthly_stats = []
        for row in raw_monthly:
            active_days = row['active_days'] or 1
            monthly_stats.append({
                'month_year': row['month_year'],
                'total_guests': row['total_guests'],
                'total_items': row['total_items'],
                'avg_guests_per_day': row['total_guests'] / active_days,
                'avg_items_per_day': row['total_items'] / active_days,
            })

        cursor.execute(
            f"SELECT COUNT(DISTINCT customer_id || '_' || checked_out_date) AS total_guests, COUNT(*) AS total_checkouts, COUNT(DISTINCT checked_out_date) AS total_days FROM loans WHERE agreement_pending = 0 {year_clause}",
            params,
        )
        summary_row = cursor.fetchone()
        if summary_row and summary_row['total_checkouts']:
            total_days = summary_row['total_days'] or 1
            analytics_summary = {
                'total_checkouts': summary_row['total_checkouts'],
                'unique_guests': summary_row['total_guests'],
                'avg_per_day': summary_row['total_guests'] / total_days,
            }

    cursor.execute("SELECT DISTINCT strftime('%Y', checked_out_date) as year FROM loans WHERE agreement_pending = 0 ORDER BY year DESC")
    years = [row['year'] for row in cursor.fetchall() if row['year']]
    conn.close()

    return render_template(
        'reports.html',
        report_data=report_data,
        report_type=report_type,
        report_title=report_title,
        year_filter=year_filter,
        month_filter=month_filter,
        date_from=date_from,
        date_to=date_to,
        years=years,
        analytics_months=analytics_months,
        analytics_summary=analytics_summary,
        daily_guests=daily_guests,
        monthly_stats=monthly_stats,
    )


@app.route('/customer_agreement/<int:customer_id>', methods=['GET', 'POST'])
@login_required
def customer_agreement(customer_id):
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute("SELECT id, name, phone, zip_code FROM customers WHERE id = ?", (customer_id,))
    customer = cursor.fetchone()
    conn.close()

    if not customer:
        flash('Customer not found.')
        return redirect(url_for('customers'))

    new_customer = request.args.get('new_customer') or request.form.get('new_customer')
    loan_ids_csv = request.args.get('loan_ids')
    loan_id = request.args.get('loan_id')
    loans = []
    due_date = None
    checkout_date = request.form.get('checkout_date', '') if request.method == 'POST' else request.args.get('checkout_date', '')
    agreement_date = request.form.get('agreement_date', '') if request.method == 'POST' else ''
    if loan_ids_csv:
        loan_ids = [int(x) for x in loan_ids_csv.split(',') if x.strip().isdigit()]
        if loan_ids:
            conn = connect_db()
            cursor = conn.cursor()
            placeholders = ",".join("?" for _ in loan_ids)
            cursor.execute(
                f"SELECT loans.id, loans.equipment_id, equipment.item_name, loans.checked_out_date, loans.due_date, loans.agreement_date FROM loans "
                f"LEFT JOIN equipment ON loans.equipment_id = equipment.equipment_id "
                f"WHERE loans.id IN ({placeholders})",
                loan_ids,
            )
            loans = cursor.fetchall()
            conn.close()
            if loans:
                due_date = loans[0]['due_date']
                if not checkout_date and 'checked_out_date' in dict(loans[0]):
                    checkout_date = loans[0]['checked_out_date']
                if not agreement_date and 'agreement_date' in dict(loans[0]):
                    agreement_date = loans[0]['agreement_date']
    elif loan_id:
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute(
            "SELECT loans.id, loans.equipment_id, equipment.item_name, loans.checked_out_date, loans.due_date, loans.agreement_date FROM loans "
            "LEFT JOIN equipment ON loans.equipment_id = equipment.equipment_id "
            "WHERE loans.id = ?",
            (loan_id,)
        )
        loan_row = cursor.fetchone()
        conn.close()
        if loan_row:
            loans = [loan_row]
            due_date = loan_row['due_date']
            if not checkout_date and 'checked_out_date' in dict(loan_row):
                checkout_date = loan_row['checked_out_date']
            if not agreement_date and 'agreement_date' in dict(loan_row):
                agreement_date = loan_row['agreement_date']

    if request.method == 'POST':
        action = request.form.get('action', 'save')
        loan_ids_csv = request.form.get('loan_ids') or loan_ids_csv
        loan_id = request.form.get('loan_id') or loan_id

        if action == 'cancel':
            loans_to_cancel = []
            if loan_ids_csv:
                loans_to_cancel = [int(x) for x in loan_ids_csv.split(',') if x.strip().isdigit()]
            elif loan_id:
                loans_to_cancel = [int(loan_id)]
            if loans_to_cancel:
                conn = connect_db()
                cursor = conn.cursor()
                cursor.execute(
                    f"SELECT equipment_id, checked_out_date FROM loans WHERE id IN ({','.join('?' for _ in loans_to_cancel)})",
                    loans_to_cancel,
                )
                cancelled_info = cursor.fetchall()
                for row in cancelled_info:
                    cursor.execute(
                        "DELETE FROM checkout_log WHERE equipment_id = ? AND checkout_date = ?",
                        (row['equipment_id'], row['checked_out_date']),
                    )
                cursor.execute(
                    f"DELETE FROM loans WHERE id IN ({','.join('?' for _ in loans_to_cancel)}) AND agreement_pending = 1",
                    loans_to_cancel,
                )
                conn.commit()
                conn.close()
            flash('Checkout cancelled and pending items removed.')
            return redirect(url_for('master_control'))

        if action == 'add_equipment':
            selected_equipment = request.form.getlist('equipment_ids')
            new_equipment_id = request.form.get('new_equipment_id', '').strip().upper()
            new_item_name = request.form.get('new_item_name', '').strip()
            checkout_date = normalize_date_input(request.form.get('checkout_date', '')) or datetime.today().date().isoformat()
            if not selected_equipment and not new_equipment_id:
                flash('Please select at least one equipment item to add or provide a new equipment entry.')
                return redirect(url_for('customer_agreement', customer_id=customer_id, loan_ids=loan_ids_csv, loan_id=loan_id, new_customer=new_customer))

            conn = connect_db()
            cursor = conn.cursor()

            if new_customer:
                new_name = request.form.get('add_equipment_new_name', '').strip()
                new_phone = request.form.get('add_equipment_new_phone', '').strip()
                new_zip = request.form.get('add_equipment_new_zip', '').strip()

                if new_name:
                    digits = normalize_phone(new_phone) if new_phone else ''
                    if new_phone and len(digits) < 10:
                        conn.close()
                        flash('Phone number must have at least 10 digits.')
                        return redirect(url_for('customer_agreement', customer_id=customer_id, loan_ids=loan_ids_csv, loan_id=loan_id, new_customer=1))
                    if digits and customer_phone_exists(digits, cursor, exclude_id=customer_id):
                        conn.close()
                        flash('A customer with this phone number already exists.')
                        return redirect(url_for('customer_agreement', customer_id=customer_id, loan_ids=loan_ids_csv, loan_id=loan_id, new_customer=1))
                    cursor.execute(
                        "UPDATE customers SET name = ?, phone = ?, zip_code = ? WHERE id = ?",
                        (new_name, format_phone(new_phone) if new_phone else '', new_zip if new_zip else '00000', customer_id),
                    )
                    conn.commit()
                    new_customer = None

            cursor.execute("SELECT zip_code FROM customers WHERE id = ?", (customer_id,))
            cust_row = cursor.fetchone()
            customer_zip = cust_row['zip_code'] if cust_row else ''

            if new_equipment_id:
                if not new_item_name:
                    conn.close()
                    flash('New equipment item name is required when adding a new equipment ID.')
                    return redirect(url_for('customer_agreement', customer_id=customer_id, loan_ids=loan_ids_csv, loan_id=loan_id, new_customer=new_customer))
                if not EQUIPMENT_ID_PATTERN.fullmatch(new_equipment_id):
                    conn.close()
                    flash('New equipment ID must be in format AA-0000.')
                    return redirect(url_for('customer_agreement', customer_id=customer_id, loan_ids=loan_ids_csv, loan_id=loan_id, new_customer=new_customer))
                cursor.execute("SELECT equipment_id FROM equipment WHERE equipment_id = ?", (new_equipment_id,))
                if cursor.fetchone():
                    conn.close()
                    flash(f'Equipment {new_equipment_id} already exists. Please select it from the available items.')
                    return redirect(url_for('customer_agreement', customer_id=customer_id, loan_ids=loan_ids_csv, loan_id=loan_id, new_customer=new_customer))
                cursor.execute(
                    "INSERT INTO equipment (equipment_id, item_name) VALUES (?, ?)",
                    (new_equipment_id, new_item_name),
                )
                selected_equipment.append(new_equipment_id)

            for equipment_id in selected_equipment:
                equipment_id = equipment_id.strip().upper()
                if not equipment_id:
                    continue
                cursor.execute(
                    "SELECT equipment_id, item_name FROM equipment WHERE equipment_id = ?",
                    (equipment_id,),
                )
                equipment_row = cursor.fetchone()
                if not equipment_row:
                    continue
                cursor.execute(
                    "SELECT id FROM loans WHERE equipment_id = ? AND returned_date IS NULL",
                    (equipment_id,),
                )
                if cursor.fetchone():
                    continue
                due_date_value = calculate_due_date(checkout_date, CHECKOUT_PERIOD_DAYS)
                cursor.execute(
                    "INSERT INTO loans (customer_id, equipment_id, checked_out_date, due_date, agreement_data, agreement_date, agreement_pending) VALUES (?, ?, ?, ?, ?, ?, 1)",
                    (customer_id, equipment_id, checkout_date, due_date_value, None, None),
                )
                new_loan_id = cursor.lastrowid
                if loan_ids_csv:
                    loan_ids_csv += f",{new_loan_id}"
                else:
                    loan_ids_csv = str(new_loan_id)
            conn.commit()
            conn.close()
            redirect_params = {'customer_id': customer_id, 'loan_ids': loan_ids_csv}
            redirect_params['checkout_date'] = checkout_date
            if new_customer:
                redirect_params['new_customer'] = new_customer
            return redirect(url_for('customer_agreement', **redirect_params))

        if action == 'save':
            if new_customer:
                new_name = request.form.get('new_name', '').strip()
                new_phone = request.form.get('new_phone', '').strip()
                new_zip = request.form.get('new_zip', '').strip()

                if new_name:
                    digits = normalize_phone(new_phone) if new_phone else ''
                    if new_phone and len(digits) < 10:
                        flash('Phone number must have at least 10 digits.')
                        return redirect(url_for('customer_agreement', customer_id=customer_id, loan_ids=loan_ids_csv, loan_id=loan_id))
                    conn = connect_db()
                    cursor = conn.cursor()
                    if digits and customer_phone_exists(digits, cursor, exclude_id=customer_id):
                        conn.close()
                        flash('A customer with this phone number already exists.')
                        return redirect(url_for('customer_agreement', customer_id=customer_id, loan_ids=loan_ids_csv, loan_id=loan_id))
                    cursor.execute(
                        "UPDATE customers SET name = ?, phone = ?, zip_code = ? WHERE id = ?",
                        (new_name, format_phone(new_phone) if new_phone else '', new_zip if new_zip else '00000', customer_id),
                    )
                    conn.commit()
                    conn.close()

            waiver_agreed = 'waiver_agreed' in request.form
            signature_agreed = 'signature_agreed' in request.form
            signature_data = request.form.get('signature_data', '')
            checkout_date = normalize_date_input(request.form.get('checkout_date', '')) or datetime.today().date().isoformat()
            due_date = calculate_due_date(checkout_date, CHECKOUT_PERIOD_DAYS) or datetime.today().date().isoformat()
            agreement_date = normalize_date_input(request.form.get('agreement_date', '')) or datetime.today().date().isoformat()

            if not waiver_agreed or not signature_agreed:
                flash('You must agree to both the waiver and digital signature acknowledgement.')
                return redirect(url_for('customer_agreement', customer_id=customer_id, loan_ids=loan_ids_csv, loan_id=loan_id))

            if not signature_data:
                flash('Please provide a digital signature.')
                return redirect(url_for('customer_agreement', customer_id=customer_id, loan_ids=loan_ids_csv, loan_id=loan_id))

            conn = connect_db()
            cursor = conn.cursor()
            updated_loans = []
            if loan_ids_csv:
                loan_ids = [int(x) for x in loan_ids_csv.split(',') if x.strip().isdigit()]
                for lid in loan_ids:
                    cursor.execute(
                        "UPDATE loans SET checked_out_date = ?, due_date = ?, agreement_data = ?, agreement_date = ?, agreement_pending = 0 WHERE id = ? AND agreement_pending = 1",
                        (checkout_date, due_date, signature_data, agreement_date, lid),
                    )
                    if cursor.rowcount:
                        updated_loans.append(lid)
            elif loan_id:
                cursor.execute(
                    "UPDATE loans SET checked_out_date = ?, due_date = ?, agreement_data = ?, agreement_date = ?, agreement_pending = 0 WHERE id = ? AND agreement_pending = 1",
                    (checkout_date, due_date, signature_data, agreement_date, loan_id),
                )
                if cursor.rowcount:
                    updated_loans.append(int(loan_id))

            if updated_loans:
                cursor.execute("SELECT zip_code FROM customers WHERE id = ?", (customer_id,))
                customer_zip = cursor.fetchone()['zip_code']
                placeholders = ','.join('?' for _ in updated_loans)
                cursor.execute(
                    f"SELECT loans.equipment_id, equipment.item_name FROM loans LEFT JOIN equipment ON loans.equipment_id = equipment.equipment_id WHERE loans.id IN ({placeholders})",
                    updated_loans,
                )
                for index, row in enumerate(cursor.fetchall()):
                    cursor.execute(
                        "INSERT INTO checkout_log (customer_zip_code, item_name, equipment_id, checkout_date, is_first_item) VALUES (?, ?, ?, ?, ?)",
                        (customer_zip, row['item_name'] or row['equipment_id'], row['equipment_id'], checkout_date, 1 if index == 0 else 0),
                    )
                cursor.execute(
                    "INSERT INTO customer_agreements (customer_id, loan_id, waiver_agreed, digital_signature_agreed, signature_data, agreed_date) VALUES (?, ?, ?, ?, ?, ?)",
                    (customer_id, updated_loans[0], 1 if waiver_agreed else 0, 1 if signature_agreed else 0, signature_data, datetime.today().date().isoformat()),
                )
            conn.commit()
            conn.close()
            flash('Customer agreement recorded successfully.')
            return redirect(url_for('master_control'))

    if not checkout_date:
        checkout_date = datetime.today().date().isoformat()
    if not due_date:
        due_date = calculate_due_date(checkout_date, CHECKOUT_PERIOD_DAYS)
    if not agreement_date:
        agreement_date = datetime.today().date().isoformat()

    return render_template(
        'customer_agreement.html',
        customer=customer,
        loan_id=loan_id,
        loan_ids=loan_ids_csv,
        due_date=due_date,
        checkout_period_days=CHECKOUT_PERIOD_DAYS,
        loans=loans,
        search_api_key=SEARCH_API_KEY or '',
        new_customer=new_customer,
        agreement_date=agreement_date,
        checkout_date=checkout_date,
    )


@app.route('/api/equipment_search')
def equipment_search():
    q = request.args.get('q', '').strip()
    try:
        limit = int(request.args.get('limit', 20))
    except Exception:
        limit = 20

    if SEARCH_API_KEY:
        provided_key = request.headers.get('X-API-KEY', '')
        if provided_key != SEARCH_API_KEY:
            return jsonify({'error': 'Unauthorized'}), 401

    client_ip = request.headers.get('X-Forwarded-For', request.remote_addr) or 'unknown'
    now = datetime.now().timestamp()
    timestamps = [t for t in search_request_log.get(client_ip, []) if now - t < SEARCH_RATE_PERIOD]
    if len(timestamps) >= SEARCH_RATE_LIMIT:
        return jsonify({'error': 'Too many requests'}), 429
    timestamps.append(now)
    search_request_log[client_ip] = timestamps

    conn = connect_db()
    cursor = conn.cursor()
    params = []
    query = (
        "SELECT equipment.equipment_id, equipment.item_name FROM equipment "
        "LEFT JOIN loans ON equipment.equipment_id = loans.equipment_id AND loans.returned_date IS NULL "
        "WHERE loans.id IS NULL "
    )
    if q:
        qpat = f"%{escape_like(q)}%"
        query += "AND (equipment.equipment_id LIKE ? ESCAPE '\\' OR equipment.item_name LIKE ? ESCAPE '\\') "
        params.extend([qpat, qpat])
    query += "ORDER BY equipment.equipment_id LIMIT ?"
    params.append(limit)
    cursor.execute(query, tuple(params))
    rows = cursor.fetchall()
    conn.close()
    results = [{"equipment_id": r["equipment_id"], "item_name": r["item_name"]} for r in rows]
    return jsonify(results)


@app.route('/agreement_view/<int:loan_id>')
@login_required
def agreement_view(loan_id):
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT customer_id FROM loans WHERE id = ? AND returned_date IS NULL AND agreement_data IS NOT NULL",
        (loan_id,),
    )
    agreement = cursor.fetchone()
    conn.close()
    if not agreement:
        flash('Agreement not found or no saved signature available.')
        return redirect(url_for('master_control'))
    return redirect(url_for('customer_agreement_view', customer_id=agreement['customer_id']))


@app.route('/customer_agreement_view/<int:customer_id>')
@login_required
def customer_agreement_view(customer_id):
    search = request.args.get('search', '').strip()
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute("SELECT id, name, phone, zip_code FROM customers WHERE id = ?", (customer_id,))
    customer = cursor.fetchone()
    if not customer:
        conn.close()
        flash('Customer not found.')
        return redirect(url_for('customers', search=search))

    cursor.execute(
        "SELECT loans.id, loans.equipment_id, equipment.item_name, loans.checked_out_date, loans.due_date, "
        "loans.agreement_date, loans.agreement_data "
        "FROM loans LEFT JOIN equipment ON loans.equipment_id = equipment.equipment_id "
        "WHERE loans.customer_id = ? AND loans.returned_date IS NULL AND loans.agreement_data IS NOT NULL "
        "ORDER BY loans.checked_out_date, loans.id",
        (customer_id,),
    )
    agreements = cursor.fetchall()
    conn.close()
    if not agreements:
        flash('No signed active agreement found for this customer.')
        return redirect(url_for('customers', search=search))

    latest_agreement = max(agreements, key=lambda loan: (loan['agreement_date'] or '', loan['id']))
    return render_template(
        'agreement_view.html',
        customer=customer,
        agreements=agreements,
        signature_data=latest_agreement['agreement_data'],
        agreement_date=latest_agreement['agreement_date'],
    )


@app.route('/inline_update', methods=['POST'])
@login_required
def inline_update():
    data = request.get_json()
    if not data:
        return jsonify({'error': 'Invalid request'}), 400

    table = data.get('table')
    row_id = data.get('row_id')
    field = data.get('field')
    value = data.get('value')

    if not all([table, row_id, field, value is not None]):
        return jsonify({'error': 'Missing required fields'}), 400

    allowed_tables = {'loans', 'customers', 'equipment'}
    if table not in allowed_tables:
        return jsonify({'error': 'Invalid table'}), 400

    allowed_fields = {
        'loans': {'checked_out_date', 'due_date', 'returned_date', 'equipment_id', 'item_name'},
        'customers': {'name', 'phone', 'zip_code'},
        'equipment': {'item_name', 'equipment_id', 'date_verified'},
    }
    if field not in allowed_fields.get(table, set()):
        return jsonify({'error': 'Invalid field'}), 400

    try:
        conn = connect_db()
        cursor = conn.cursor()
        if table == 'loans':
            cursor.execute(f"UPDATE loans SET {field} = ? WHERE id = ?", (value, int(row_id)))
        elif table == 'customers':
            if field == 'phone':
                digits = normalize_phone(value)
                if len(digits) < 10:
                    conn.close()
                    return jsonify({'error': 'Phone number must have at least 10 digits.'}), 400
                if customer_phone_exists(digits, cursor, exclude_id=int(row_id)):
                    conn.close()
                    return jsonify({'error': 'Phone number already exists for another customer.'}), 400
                value = format_phone(value)
            cursor.execute(f"UPDATE customers SET {field} = ? WHERE id = ?", (value, int(row_id)))
        elif table == 'equipment':
            if field == 'equipment_id':
                cursor.execute("UPDATE loans SET equipment_id = ? WHERE equipment_id = ?", (value, row_id))
                cursor.execute("UPDATE checkout_log SET equipment_id = ? WHERE equipment_id = ?", (value, row_id))
                cursor.execute("UPDATE deleted_items_log SET equipment_id = ? WHERE equipment_id = ?", (value, row_id))
                cursor.execute("UPDATE equipment SET equipment_id = ? WHERE equipment_id = ?", (value, row_id))
            else:
                cursor.execute(f"UPDATE equipment SET {field} = ? WHERE equipment_id = ?", (value, row_id))
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        log.exception("Inline update error")
        return jsonify({'error': 'An internal error occurred.'}), 500


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--blank', action='store_true', help='Delete existing database and create a fresh blank one')
    parser.add_argument('--port', type=int, default=5000)
    args = parser.parse_args()

    if args.blank:
        blank_db()
    else:
        init_db()

    admin_user = os.environ.get('ADMIN_USERNAME')
    admin_pass = os.environ.get('ADMIN_PASSWORD')
    if admin_user and admin_pass:
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM users WHERE username = ?", (admin_user,))
        if not cursor.fetchone():
            cursor.execute(
                "INSERT INTO users (username, password_hash, is_admin) VALUES (?, ?, 1)",
                (admin_user, generate_password_hash(admin_pass)),
            )
            conn.commit()
            log.info("Auto-created admin user: %s", admin_user)
        conn.close()

    debug = os.environ.get('FLASK_DEBUG', 'false').lower() == 'true'
    ssl_certfile = os.environ.get('SSL_CERTFILE')
    ssl_keyfile = os.environ.get('SSL_KEYFILE')
    ssl_context = None
    if ssl_certfile and ssl_keyfile:
        ssl_context = (ssl_certfile, ssl_keyfile)
        log.info("TLS enabled with cert=%s", ssl_certfile)

    log.info("Starting DME Checkout app on port %d (debug=%s)", args.port, debug)

    if getattr(sys, 'frozen', False):
        proto = 'https' if ssl_context else 'http'
        url = f'{proto}://localhost:{args.port}'
        threading.Timer(1.5, lambda: webbrowser.open(url)).start()
        log.info("Browser will open to %s", url)

    app.run(debug=debug, host='0.0.0.0', port=args.port, ssl_context=ssl_context)
